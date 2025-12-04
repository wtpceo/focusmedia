#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
엑셀 파일을 읽어서 data.json으로 변환하는 스크립트
엘리베이터TV 설치리스트 데이터를 지도 웹앱용 JSON으로 변환
"""

import json
from openpyxl import load_workbook

def clean_number(value):
    """숫자 값 정리"""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    try:
        return int(str(value).replace(',', '').strip())
    except:
        return 0

def clean_text(value):
    """텍스트 정리"""
    if value is None:
        return ''
    return str(value).strip()

def main():
    input_file = '/Users/wtpceo/Desktop/01.위플 프로젝트/01.개발/08.homepages_1/03.focus/엘리베이터TV 설치리스트(외부용)_ver 251124.xlsx'
    output_file = '/Users/wtpceo/Desktop/01.위플 프로젝트/01.개발/08.homepages_1/03.focus/data.json'

    print(f"엑셀 파일 로딩 중: {input_file}")
    wb = load_workbook(input_file, read_only=True, data_only=True)
    ws = wb.active

    # 모든 행을 리스트로 변환
    rows = list(ws.iter_rows(values_only=True))
    print(f"총 {len(rows)}행 읽음")

    # 헤더 찾기 (단지명이 있는 행)
    header_row = None
    for i, row in enumerate(rows):
        if row and any('단지명' in str(cell) if cell else False for cell in row):
            header_row = i
            break

    if header_row is None:
        print("헤더를 찾을 수 없습니다.")
        return

    headers = [str(h).strip() if h else '' for h in rows[header_row]]
    print(f"헤더 행: {header_row}")
    print(f"헤더: {headers[:20]}")  # 처음 20개만 출력

    # 컬럼 인덱스 찾기
    col_indices = {}
    for i, h in enumerate(headers):
        h_clean = h.replace('\n', ' ')
        if '단지명' in h_clean:
            col_indices['name'] = i
        elif '단지코드' in h_clean:
            col_indices['code'] = i
        elif h_clean == '도시':
            col_indices['city'] = i
        elif h_clean == '구':
            col_indices['gu'] = i
        elif '동(법정동)' in h_clean or h_clean == '동':
            col_indices['dong'] = i
        elif '주소(도로명)' in h_clean:
            col_indices['address'] = i
        elif '주소(지번)' in h_clean:
            col_indices['address_jibun'] = i
        elif '건물유형' in h_clean:
            col_indices['building_type'] = i
        elif '준공연도' in h_clean:
            col_indices['year'] = i
        elif '건물층수' in h_clean:
            col_indices['floors'] = i
        elif '기준평형' in h_clean:
            col_indices['area'] = i
        elif '총 세대수' in h_clean or '총세대수' in h_clean:
            col_indices['households'] = i
        elif '총 인구수' in h_clean or '총인구수' in h_clean:
            col_indices['population'] = i
        elif '판매등급' in h_clean:
            col_indices['grade'] = i
        elif '판매수량' in h_clean:
            col_indices['quantity'] = i
        elif '4주 금액' in h_clean or '4주금액' in h_clean:
            col_indices['price_4w'] = i
        elif '대당단가' in h_clean:
            col_indices['unit_price'] = i
        elif '최초 설치일자' in h_clean or '최초설치일자' in h_clean:
            col_indices['install_date'] = i

    print(f"찾은 컬럼 인덱스: {col_indices}")

    locations = []

    # 데이터 행 처리
    for row in rows[header_row + 1:]:
        if not row or len(row) <= col_indices.get('name', 5):
            continue

        name = clean_text(row[col_indices.get('name', 5)])
        if not name:
            continue

        # 주소 가져오기 (도로명 우선, 없으면 지번)
        address = clean_text(row[col_indices.get('address', 11)])
        if not address:
            address = clean_text(row[col_indices.get('address_jibun', 10)])

        if not address:
            continue

        location = {
            'name': name,
            'code': clean_text(row[col_indices.get('code', 6)]),
            'city': clean_text(row[col_indices.get('city', 7)]),
            'gu': clean_text(row[col_indices.get('gu', 8)]),
            'dong': clean_text(row[col_indices.get('dong', 9)]),
            'address': address,
            'building_type': clean_text(row[col_indices.get('building_type', 12)]),
            'year': clean_number(row[col_indices.get('year', 13)]),
            'floors': clean_number(row[col_indices.get('floors', 14)]),
            'area': clean_number(row[col_indices.get('area', 15)]),
            'households': clean_number(row[col_indices.get('households', 16)]),
            'population': clean_number(row[col_indices.get('population', 17)]),
            'grade': clean_text(row[col_indices.get('grade', 19)]),
            'quantity': clean_number(row[col_indices.get('quantity', 20)]),
            'unit_price': clean_number(row[col_indices.get('unit_price', 21)]),
            'price_4w': clean_number(row[col_indices.get('price_4w', 22)]),
            'install_date': clean_text(row[col_indices.get('install_date', 18)])
        }

        locations.append(location)

    wb.close()

    # JSON 파일 저장
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(locations, f, ensure_ascii=False, indent=2)

    print(f"\n총 {len(locations)}개 데이터 변환 완료")
    print(f"저장 위치: {output_file}")

    # 샘플 출력
    if locations:
        print("\n샘플 데이터 (첫 번째):")
        print(json.dumps(locations[0], ensure_ascii=False, indent=2))
        print("\n샘플 데이터 (마지막):")
        print(json.dumps(locations[-1], ensure_ascii=False, indent=2))

if __name__ == '__main__':
    main()
